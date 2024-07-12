# import the necessary libraries
import openpyxl
import time
import os

# OPTION ONE
# def get_current_price(sheet, product_name):
#     for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row, values_only=False):
#         if row[0].value == product_name:
#             return row[1].value
#     return None

# def update_price_and_mark_updated(file_path, product_name, new_price):
#     # Load the workbook and select the active worksheet
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active

#     current_price = get_current_price(sheet, product_name)
#     if current_price is not None:
#         print(f"The current price of {product_name} is {current_price}")
#         new_price = float(input(f"Enter a new price for {product_name}: "))

#         # Update the price and mark as updated
#         for row in sheet.iter_rows(min_row=2, max_col=5, values_only=False):
#             if row[0].value == product_name:
#                 row[1].value = new_price  # Update the price in column B
#                 total = round(new_price * row[2].value, 2)  # Calculate the total for column D
#                 row[3].value = total  # Update the total in column D
#                 row[4].value = "UPDATED"  # Mark as updated in column E
#             else:
#                 total = round(row[1].value * row[2].value, 2)  # Calculate the total for column D
#                 row[3].value = total  # Update the total in column D

#         # Save the updated workbook to a new file
#         new_file_path = 'products_revised.xlsx'
#         workbook.save(new_file_path)
#         print(f"Updated file saved as {new_file_path}")
#     else:
#         print(f"The product '{product_name}' was not found in the file.")

# def main():
#     # Define the path to the original Excel file
#     file_path = 'produceSales.xlsx'

#     # Prompt the user for input
#     product_name = input("Enter the product name: ")

#     # Update the price and mark as updated in the Excel file
#     update_price_and_mark_updated(file_path, product_name, 0)

# if __name__ == "__main__":
#     main()


# OPTION TWO ------------


def load_products(file_path: str) -> dict[str, float]:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    products = {}
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        products[row[0]] = row[1]
    return products


def update_price_and_mark_updated(
    file_path: str, product_name: str, new_price: float
) -> None:
    products = load_products(file_path)
    if product_name not in products:
        raise ValueError(f"The product '{product_name}' was not found in the file.")

    current_price = products[product_name]
    print(f"The current price of {product_name} is {current_price}")
    new_price = float(input(f"Enter a new price for {product_name}: "))

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    quantities = {}
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        quantities[row[0]] = row[1]

    for row in sheet.iter_rows(min_row=2, max_col=5, values_only=False):
        if row[0].value == product_name:
            row[1].value = new_price
            total = round(new_price * quantities[row[0].value], 2)
            row[3].value = total
            row[4].value = "UPDATED"
        elif row[0].value in quantities:
            total = round(row[1].value * quantities[row[0].value], 2)
            row[3].value = total

    new_file_path = "products_revised.xlsx"
    workbook.save(new_file_path)
    print(f"Updated file saved as {new_file_path}")


def main() -> None:
    file_path = "produceSales.xlsx"
    product_name = input("Enter the product name: ")
    try:
        update_price_and_mark_updated(file_path, product_name, 0)
    except ValueError as e:
        print(str(e))


if __name__ == "__main__":
    main()

print("Done!")
time.sleep(3)

# Clear the terminal screen
os.system("clear")
