
import time
from openpyxl import Workbook

# Prompt user to enter a number
num = int(input("Enter a number: "))

# Measure the elapsed time
start_time = time.time()

# Create a new Excel file
wb = Workbook()
sheet = wb.active
sheet.title = "multiplied"

# Fill the first row and column in one go
for i in range(1, num + 2):
    sheet.cell(row=1, column=i, value=i - 1)
    sheet.cell(row=i, column=1, value=i - 1)

# Prepare lists for row and column headers
row_headers = [i - 1 for i in range(1, num + 2)]
col_headers = [i - 1 for i in range(1, num + 2)]

# Perform multiplication for the rest of the cells
for row in range(2, num + 2):
    for col in range(2, num + 2):
        sheet.cell(row=row, column=col, value=row_headers[row - 1] * col_headers[col - 1])

# Save the Excel file
wb.save("my_multiplication.xlsx")

# Calculate and display the elapsed time
elapsed_time = time.time() - start_time
print(f"Elapsed time: {elapsed_time:.2f} seconds")

print("Done!")
